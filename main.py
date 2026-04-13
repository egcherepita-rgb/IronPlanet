import io
import os
import re
import time
import json
import threading
from uuid import uuid4
from collections import OrderedDict
from typing import Dict, List, Optional, Tuple

import fitz  # PyMuPDF
from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.responses import Response, HTMLResponse, FileResponse
from fastapi.staticfiles import StaticFiles
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


app = FastAPI(
    title="PDF → XLSX (IronPlanet)",
    version="1.0.0",
)

# Static files are optional
if os.path.isdir("static"):
    app.mount("/static", StaticFiles(directory="static"), name="static")


# -------------------------
# Regex
# -------------------------
RX_SIZE = re.compile(r"\b\d{2,}[xх×]\d{2,}(?:[xх×]\d{1,})?\b", re.IGNORECASE)
RX_MM = re.compile(r"мм", re.IGNORECASE)
RX_WEIGHT = re.compile(r"\b\d+(?:[.,]\d+)?\s*кг\.?\b", re.IGNORECASE)
RX_MONEY_LINE = re.compile(r"^\d+(?:[ \u00a0]\d{3})*(?:[.,]\d+)?\s*₽$")
RX_INT = re.compile(r"^\d+$")
RX_ANY_RUB = re.compile(r"₽")
RX_DIMS_ANYWHERE = re.compile(
    r"\s*\d{1,4}[xх×]\d{1,4}(?:[xх×]\d{1,5})?\s*мм\.?\s*",
    re.IGNORECASE,
)

# Inline: ... 450 ₽ 2 900 ₽
RX_PRICE_QTY_SUM = re.compile(
    r"(?P<price>\d+(?:[ \u00a0]\d{3})*(?:[.,]\d+)?)\s*₽\s+"
    r"(?P<qty>\d{1,4})\s+"
    r"(?P<sum>\d+(?:[ \u00a0]\d{3})*(?:[.,]\d+)?)\s*₽",
    re.IGNORECASE,
)


# -------------------------
# Helpers
# -------------------------
def normalize_space(s: str) -> str:
    s = (s or "").replace("\u00a0", " ")
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def normalize_key(name: str) -> str:
    s = normalize_space(name).lower()
    s = s.replace("×", "x").replace("х", "x")
    s = RX_DIMS_ANYWHERE.sub(" ", s)
    return normalize_space(s)


def strip_dims_anywhere(name: str) -> str:
    return normalize_space(RX_DIMS_ANYWHERE.sub(" ", normalize_space(name)))


def money_to_number(text: str) -> int:
    s = normalize_space(text).replace("₽", "").replace(" ", "").replace(",", ".")
    try:
        return int(round(float(s)))
    except Exception:
        return -1


# -------------------------
# Mappings
# -------------------------
def load_source_article_map() -> Tuple[Dict[str, str], str]:
    path = os.getenv("SOURCE_ART_XLSX_PATH", "Art.xlsx")
    if not os.path.exists(path):
        return {}, f"file_not_found:{path}"

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[wb.sheetnames[0]]
    except Exception as e:
        return {}, f"cannot_open:{e}"

    header = [normalize_space(ws.cell(1, c).value or "") for c in range(1, ws.max_column + 1)]
    товар_col = 1
    арт_col = 2
    for idx, h in enumerate(header, start=1):
        if h.lower() == "товар":
            товар_col = idx
        if h.lower() == "артикул":
            арт_col = idx

    mapping: Dict[str, str] = {}
    for r in range(2, ws.max_row + 1):
        src_name = ws.cell(r, товар_col).value
        article = ws.cell(r, арт_col).value
        if not src_name or not article:
            continue
        mapping[normalize_key(str(src_name))] = normalize_space(str(article))
    return mapping, "ok"


def load_ironplanet_map() -> Tuple[Dict[str, Dict[str, str]], str]:
    path = os.getenv("IRONPLANET_XLSX_PATH", "IRONPLANET.xlsx")
    if not os.path.exists(path):
        return {}, f"file_not_found:{path}"

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[wb.sheetnames[0]]
    except Exception as e:
        return {}, f"cannot_open:{e}"

    header = [normalize_space(ws.cell(1, c).value or "") for c in range(1, ws.max_column + 1)]
    article_col = 1
    name_col = 2
    kodtov_col = 3
    for idx, h in enumerate(header, start=1):
        low = h.lower()
        if low == "артикул":
            article_col = idx
        elif low == "наименование":
            name_col = idx
        elif low == "kodtov":
            kodtov_col = idx

    mapping: Dict[str, Dict[str, str]] = {}
    for r in range(2, ws.max_row + 1):
        article = ws.cell(r, article_col).value
        custom_name = ws.cell(r, name_col).value
        kodtov = ws.cell(r, kodtov_col).value
        if not article:
            continue
        article_s = normalize_space(str(article))
        mapping[article_s] = {
            "article": article_s,
            "name": normalize_space(str(custom_name or "")),
            "kodtov": normalize_space(str(kodtov or "")),
        }
    return mapping, "ok"


SOURCE_ARTICLE_MAP, SOURCE_ARTICLE_MAP_STATUS = load_source_article_map()
IRONPLANET_MAP, IRONPLANET_MAP_STATUS = load_ironplanet_map()


# -------------------------
# Counter
# -------------------------
from threading import Lock

COUNTER_FILE = os.getenv("COUNTER_FILE", "conversions.count")
_counter_lock = Lock()


def _read_counter() -> int:
    try:
        with open(COUNTER_FILE, "r", encoding="utf-8") as f:
            return int((f.read() or "0").strip())
    except Exception:
        return 0


def _write_counter(v: int) -> None:
    tmp = COUNTER_FILE + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        f.write(str(v))
    os.replace(tmp, COUNTER_FILE)


def increment_counter() -> int:
    with _counter_lock:
        v = _read_counter() + 1
        _write_counter(v)
        return v


def get_counter() -> int:
    return _read_counter()


# -------------------------
# Async jobs storage
# -------------------------
JOB_DIR = os.getenv("JOB_DIR", "/tmp/pdf2xlsx_ironplanet_jobs")
JOB_TTL_SEC = int(os.getenv("JOB_TTL_SEC", str(30 * 60)))


def _ensure_job_dir() -> None:
    os.makedirs(JOB_DIR, exist_ok=True)


def _job_json_path(job_id: str) -> str:
    return os.path.join(JOB_DIR, f"{job_id}.json")


def _job_result_path(job_id: str) -> str:
    return os.path.join(JOB_DIR, f"{job_id}.xlsx")


def _write_json_atomic(path: str, data: dict) -> None:
    _ensure_job_dir()
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False)
    os.replace(tmp, path)


def _read_json(path: str) -> Optional[dict]:
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None


def _cleanup_jobs() -> None:
    _ensure_job_dir()
    now = time.time()
    for name in os.listdir(JOB_DIR):
        if not (name.endswith(".json") or name.endswith(".xlsx")):
            continue
        p = os.path.join(JOB_DIR, name)
        try:
            if now - os.path.getmtime(p) > JOB_TTL_SEC:
                os.remove(p)
        except Exception:
            pass


def _set_job(job_id: str, **kwargs) -> None:
    _cleanup_jobs()
    path = _job_json_path(job_id)
    data = _read_json(path) or {}
    data.update(kwargs)
    _write_json_atomic(path, data)


def _get_job(job_id: str) -> Optional[dict]:
    _cleanup_jobs()
    return _read_json(_job_json_path(job_id))


def _set_job_result(job_id: str, xlsx_bytes: bytes) -> None:
    _cleanup_jobs()
    _ensure_job_dir()
    p = _job_result_path(job_id)
    tmp = p + ".tmp"
    with open(tmp, "wb") as f:
        f.write(xlsx_bytes)
    os.replace(tmp, p)


def _get_job_result(job_id: str) -> Optional[bytes]:
    p = _job_result_path(job_id)
    try:
        with open(p, "rb") as f:
            return f.read()
    except Exception:
        return None


# -------------------------
# PDF parsing helpers
# -------------------------
def is_noise(line: str) -> bool:
    low = (line or "").strip().lower()
    return (
        not low
        or low.startswith("страница:")
        or low.startswith("ваш проект")
        or "проект создан" in low
        or "развертка стены" in low
        or "стоимость проекта" in low
    )


def is_totals_block(line: str) -> bool:
    low = (line or "").strip().lower()
    return (
        low.startswith("общий вес")
        or low.startswith("максимальный габарит заказа")
        or low.startswith("адрес:")
        or low.startswith("телефон:")
        or low.startswith("email")
    )


def is_project_total_only(line: str, prev_line: str = "") -> bool:
    if not RX_MONEY_LINE.fullmatch(normalize_space(line)):
        return False
    prev = normalize_space(prev_line).lower()
    if "стоимость проекта" in prev:
        return True
    return money_to_number(line) >= 10000


def is_header_token(line: str) -> bool:
    low = normalize_space(line).lower().replace("–", "-").replace("—", "-")
    return low in {"фото", "товар", "габариты", "вес", "цена за шт", "кол-во", "сумма"}


def looks_like_dim_or_weight(line: str) -> bool:
    return bool(RX_WEIGHT.search(line) or (RX_SIZE.search(line) and RX_MM.search(line)))


def looks_like_money_or_qty(line: str) -> bool:
    return bool(RX_MONEY_LINE.fullmatch(line) or RX_INT.fullmatch(line))


def clean_name_from_buffer(buf: List[str]) -> str:
    filtered: List[str] = []
    for ln in buf:
        if is_noise(ln) or is_header_token(ln) or is_totals_block(ln):
            continue
        filtered.append(ln)

    while filtered and (looks_like_dim_or_weight(filtered[-1]) or looks_like_money_or_qty(filtered[-1])):
        filtered.pop()

    name = normalize_space(" ".join(filtered))
    name = re.sub(r"^Фото\s*", "", name, flags=re.IGNORECASE).strip()
    name = re.sub(r"^Товар\s*", "", name, flags=re.IGNORECASE).strip()
    return strip_dims_anywhere(name)


def _finalize_item(
    ordered: "OrderedDict[str, Dict[str, object]]",
    stats: Dict,
    name: str,
    qty: int,
    price: int,
    total_sum: int,
    anchor_type: str,
) -> None:
    if not name or not (0 <= qty <= 500):
        return
    if qty == 0:
        return

    key = normalize_key(name)
    total_sum = total_sum if total_sum >= 0 else price * qty if price >= 0 else 0
    if key not in ordered:
        ordered[key] = {
            "source_name": name,
            "qty": qty,
            "price": max(price, 0),
            "sum": max(total_sum, 0),
        }
    else:
        item = ordered[key]
        item["qty"] = int(item["qty"]) + qty
        item["sum"] = int(item["sum"]) + max(total_sum, 0)
        if price >= 0:
            item["price"] = price

    stats["items_found"] += 1
    if anchor_type == "inline":
        stats["anchors_inline"] += 1
    else:
        stats["anchors_multiline"] += 1


def parse_items(pdf_bytes: bytes) -> Tuple[List[Dict[str, object]], Dict]:
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    total_pages = doc.page_count

    ordered: "OrderedDict[str, Dict[str, object]]" = OrderedDict()
    buf: List[str] = []

    stats = {
        "pages": 0,
        "total_pages": total_pages,
        "processed_pages": 0,
        "items_found": 0,
        "anchors_inline": 0,
        "anchors_multiline": 0,
        "source_article_map_size": len(SOURCE_ARTICLE_MAP),
        "source_article_map_status": SOURCE_ARTICLE_MAP_STATUS,
        "ironplanet_map_size": len(IRONPLANET_MAP),
        "ironplanet_map_status": IRONPLANET_MAP_STATUS,
    }

    for page in doc:
        stats["pages"] += 1
        stats["processed_pages"] += 1

        txt = page.get_text("text") or ""
        if "₽" not in txt:
            continue

        lines = [normalize_space(x) for x in txt.splitlines()]
        lines = [x for x in lines if x]
        if not lines:
            continue

        in_totals = False
        buf.clear()
        i = 0
        while i < len(lines):
            line = lines[i]
            prev = lines[i - 1] if i > 0 else ""

            if is_noise(line) or is_header_token(line):
                i += 1
                continue

            if is_project_total_only(line, prev) or is_totals_block(line):
                in_totals = True
                buf.clear()
                i += 1
                continue

            if in_totals:
                i += 1
                continue

            # A) inline: ... price ₽ qty sum ₽
            m = RX_PRICE_QTY_SUM.search(line)
            if m:
                name = clean_name_from_buffer(buf)
                buf.clear()
                price = money_to_number(m.group("price"))
                qty = int(m.group("qty"))
                total_sum = money_to_number(m.group("sum"))
                _finalize_item(ordered, stats, name, qty, price, total_sum, "inline")
                i += 1
                continue

            # B) embedded price in product line, then qty, then sum
            if RX_ANY_RUB.search(line):
                if i + 2 < len(lines) and RX_INT.fullmatch(lines[i + 1]) and RX_MONEY_LINE.fullmatch(lines[i + 2]):
                    price_match = re.search(r"(\d+(?:[ \u00a0]\d{3})*(?:[.,]\d+)?)\s*₽", line)
                    price = money_to_number(price_match.group(1)) if price_match else -1
                    qty = int(lines[i + 1])
                    total_sum = money_to_number(lines[i + 2])
                    name = clean_name_from_buffer(buf + [re.sub(r"\d+(?:[ \u00a0]\d{3})*(?:[.,]\d+)?\s*₽.*$", "", line).strip()])
                    buf.clear()
                    _finalize_item(ordered, stats, name, qty, price, total_sum, "multiline")
                    i += 3
                    continue

            # C) multiline: price ₽ -> qty -> sum ₽
            if RX_MONEY_LINE.fullmatch(line):
                end = min(len(lines), i + 8)
                qty_idx = None
                for j in range(i + 1, end):
                    if RX_INT.fullmatch(lines[j]):
                        q = int(lines[j])
                        if 0 <= q <= 500:
                            qty_idx = j
                            break

                if qty_idx is None:
                    buf.append(line)
                    i += 1
                    continue

                sum_idx = None
                for j in range(qty_idx + 1, end):
                    if RX_MONEY_LINE.fullmatch(lines[j]):
                        sum_idx = j
                        break

                if sum_idx is None:
                    buf.append(line)
                    i += 1
                    continue

                name = clean_name_from_buffer(buf)
                buf.clear()
                price = money_to_number(line)
                qty = int(lines[qty_idx])
                total_sum = money_to_number(lines[sum_idx])
                _finalize_item(ordered, stats, name, qty, price, total_sum, "multiline")
                i = sum_idx + 1
                continue

            buf.append(line)
            i += 1

    return list(ordered.values()), stats


# -------------------------
# Transform to output rows
# -------------------------
def convert_to_output_rows(parsed_rows: List[Dict[str, object]]) -> Tuple[List[Dict[str, object]], Dict[str, int]]:
    out: List[Dict[str, object]] = []
    matched_source = 0
    matched_ironplanet = 0

    for item in parsed_rows:
        source_name = str(item["source_name"])
        qty = int(item["qty"])
        price = int(item["price"])
        total_sum = int(item["sum"])

        article = SOURCE_ARTICLE_MAP.get(normalize_key(source_name), "")
        if article:
            matched_source += 1

        iron = IRONPLANET_MAP.get(article)
        if iron:
            matched_ironplanet += 1

        out.append({
            "Артикул": article,
            "Наименование": (iron or {}).get("name") or source_name,
            "kodtov": (iron or {}).get("kodtov") or "",
            "kolvo": qty,
            "CenaR": price,
            "Сумма": total_sum if total_sum > 0 else price * qty,
        })

    return out, {
        "matched_source": matched_source,
        "matched_ironplanet": matched_ironplanet,
        "total_rows": len(out),
    }


# -------------------------
# XLSX output
# -------------------------
def make_xlsx(rows: List[Dict[str, object]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Заказ"

    headers = ["Артикул", "Наименование", "kodtov", "kolvo", "CenaR", "Сумма"]
    ws.append(headers)

    for row in rows:
        ws.append([row[h] for h in headers])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    widths = {
        "A": 18,
        "B": 70,
        "C": 14,
        "D": 10,
        "E": 12,
        "F": 14,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row in ws.iter_rows(min_row=2, min_col=1, max_col=6):
        row[0].alignment = Alignment(vertical="center")
        row[1].alignment = Alignment(wrap_text=True, vertical="center")
        for c in row[2:]:
            c.alignment = center

    for col in ["D", "E", "F"]:
        for cell in ws[col][1:]:
            cell.number_format = '0'

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# -------------------------
# UI
# -------------------------
HOME_HTML = """<!doctype html>
<html lang="ru">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>PDF → XLSX • IronPlanet</title>
  <style>
    :root { --bg:#0b0f17; --card:#121a2a; --text:#e9eefc; --muted:#a8b3d6; --border:rgba(255,255,255,.08); --btn:#4f7cff; }
    html, body { height: 100%; overflow: hidden; }
    *, *::before, *::after { box-sizing: border-box; }
    body { margin:0; font-family: system-ui, -apple-system, Segoe UI, Roboto, Arial, sans-serif; background: radial-gradient(1200px 600px at 20% 10%, #18234a 0%, var(--bg) 55%); color: var(--text); }
    .hero { position: fixed; top: 22px; left: 0; width: 100%; text-align: center; z-index: 5; }
    .hero-title { font-weight: 900; letter-spacing: .6px; font-size: clamp(28px, 4vw, 48px); margin: 0; }
    .hero-sub { margin-top: 10px; font-size: 18px; color: var(--muted); }
    .wrap { height: 100vh; display:flex; align-items:center; justify-content:center; padding: 28px; padding-top: 170px; }
    .card { width:min(920px, 100%); background: rgba(18,26,42,.92); border: 1px solid var(--border); border-radius: 18px; padding: 22px; box-shadow: 0 18px 60px rgba(0,0,0,.45); }
    .top { display:flex; gap:14px; align-items:center; justify-content:space-between; flex-wrap:wrap; }
    h1 { margin:0; font-size: 28px; }
    .hint { margin: 8px 0 0; color: var(--muted); font-size: 14px; }
    .badge { font-size: 12px; color: var(--muted); border: 1px solid var(--border); padding: 6px 10px; border-radius: 999px; }
    .row { margin-top: 18px; display:flex; gap: 12px; align-items:center; justify-content:center; flex-wrap:wrap; width:100%; }
    .file { display:flex; align-items:center; justify-content:center; gap:10px; padding: 10px 12px; border: 1px dashed var(--border); border-radius: 14px; background: rgba(255,255,255,.02); }
    button { padding: 10px 14px; border: 0; border-radius: 14px; cursor: pointer; font-weight: 800; background: var(--btn); color: #0b1020; }
    button:disabled { opacity: .55; cursor:not-allowed; }
    .status { margin-top: 14px; font-size: 14px; color: var(--muted); white-space: pre-wrap; text-align:center; }
    .status.ok { color: #79ffa8; }
    .status.err { color: #ff7b8a; }
    .cols { margin-top: 18px; padding: 14px; border-radius: 14px; border:1px solid var(--border); background: rgba(255,255,255,.02); }
    .cols strong { display:block; margin-bottom: 8px; }
    .colsgrid { display:grid; grid-template-columns: repeat(6, minmax(0, 1fr)); gap:8px; }
    .pill { text-align:center; padding: 8px 10px; border-radius: 999px; background: rgba(255,255,255,.05); font-size:13px; }
    .corner { position: fixed; right: 12px; bottom: 10px; font-size: 12px; color: var(--muted); opacity: .9; }
    .footer-note { position: fixed; left: 50%; bottom: 10px; transform: translateX(-50%); font-size: 13px; color: var(--muted); opacity: .9; text-align:center; padding: 0 12px; }
    @media (max-width: 860px) { .colsgrid { grid-template-columns: repeat(2, minmax(0, 1fr)); } .wrap { padding-top: 150px; } }
  </style>
</head>
<body>
  <div class="hero">
    <div class="hero-title">ПЛАНЕТА ЖЕЛЕЗЯКА</div>
    <div class="hero-sub">PDF → XLSX</div>
  </div>

  <div class="wrap">
    <div class="card">
      <div class="top">
        <div>
          <h1>Конвертация спецификации PDF в XLSX</h1>
         
        
      <div class="row">
        <div class="file"><input id="pdf" type="file" accept="application/pdf,.pdf" /></div>
        <button id="btn" disabled>Скачать XLSX</button>
      </div>

    
      <div id="status" class="status"></div>
    </div>
  </div>

  <div class="corner" id="counter">…</div>
  <div class="footer-note">Программа создана командой ПРОМЕТ</div>

  <script>
    const input = document.getElementById('pdf');
    const btn = document.getElementById('btn');
    const statusEl = document.getElementById('status');

    function ok(msg){ statusEl.className='status ok'; statusEl.textContent=msg; }
    function err(msg){ statusEl.className='status err'; statusEl.textContent=msg; }
    function neutral(msg){ statusEl.className='status'; statusEl.textContent=msg||''; }

    async function loadCounter(){
      try {
        const r = await fetch('/stats');
        const j = await r.json();
        document.getElementById('counter').textContent = String(j.conversions || 0);
      } catch(e) {}
    }
    loadCounter();

    input.addEventListener('change', () => {
      const f = input.files && input.files[0];
      btn.disabled = !f;
      neutral(f ? ('Выбран файл: ' + f.name) : '');
    });

    btn.addEventListener('click', async () => {
      const f = input.files && input.files[0];
      if (!f) return;
      btn.disabled = true;
      const start = Date.now();

      let timer = setInterval(() => {
        const sec = Math.floor((Date.now() - start) / 1000);
        neutral('Обработка… прошло ' + sec + ' сек');
      }, 500);

      try {
        const fd = new FormData();
        fd.append('file', f);

        const r = await fetch('/extract_async', { method: 'POST', body: fd });
        const data = await r.json();
        if (!r.ok) throw new Error(data.detail || ('HTTP ' + r.status));
        const job_id = data.job_id;

        while (true) {
          const s = await fetch('/job/' + job_id);
          const j = await s.json();
          if (!s.ok) throw new Error(j.detail || ('HTTP ' + s.status));

          let msg = (j.message || 'Обработка…');
          if (j.total_pages && j.processed_pages) {
            msg += ' • страниц: ' + j.processed_pages + '/' + j.total_pages;
          }
          neutral(msg);

          if (j.status === 'done') {
            const dl = await fetch('/job/' + job_id + '/download');
            if (!dl.ok) throw new Error('Не удалось скачать результат');
            const blob = await dl.blob();
            const filename = j.filename || ((f.name || 'items.pdf').replace(/\.pdf$/i, '') + '.xlsx');
            const url = URL.createObjectURL(blob);
            const a = document.createElement('a');
            a.href = url;
            a.download = filename;
            document.body.appendChild(a);
            a.click();
            a.remove();
            URL.revokeObjectURL(url);
            ok('Готово! Файл скачан: ' + filename);
            loadCounter();
            break;
          }

          if (j.status === 'error') throw new Error(j.message || 'Ошибка обработки');
          await new Promise(res => setTimeout(res, 600));
        }
      } catch (e) {
        err('Ошибка: ' + String(e.message || e));
      } finally {
        clearInterval(timer);
        btn.disabled = !(input.files && input.files[0]);
      }
    });
  </script>
</body>
</html>
"""


# -------------------------
# Endpoints
# -------------------------
@app.get("/stats")
def stats():
    return {"conversions": get_counter()}


@app.get("/health")
def health():
    try:
        _ensure_job_dir()
        job_files = len([x for x in os.listdir(JOB_DIR) if x.endswith(".json")])
    except Exception:
        job_files = -1

    return {
        "status": "ok",
        "source_article_map_size": len(SOURCE_ARTICLE_MAP),
        "source_article_map_status": SOURCE_ARTICLE_MAP_STATUS,
        "ironplanet_map_size": len(IRONPLANET_MAP),
        "ironplanet_map_status": IRONPLANET_MAP_STATUS,
        "conversions": get_counter(),
        "job_dir": JOB_DIR,
        "job_files": job_files,
    }


@app.api_route("/", methods=["GET", "HEAD"], response_class=HTMLResponse)
def home():
    return HOME_HTML


@app.post("/extract")
async def extract(file: UploadFile = File(...)):
    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="Загрузите PDF файл (.pdf).")

    pdf_bytes = await file.read()
    try:
        parsed_rows, stats_ = parse_items(pdf_bytes)
        output_rows, map_stats = convert_to_output_rows(parsed_rows)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Не удалось распарсить PDF: {e}")

    if not output_rows:
        raise HTTPException(status_code=422, detail=f"Не удалось найти позиции. debug={stats_}")

    xlsx_bytes = make_xlsx(output_rows)
    increment_counter()
    return Response(
        content=xlsx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="items.xlsx"'},
    )


@app.post("/extract_async")
async def extract_async(file: UploadFile = File(...)):
    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="Загрузите PDF файл (.pdf).")

    pdf_bytes = await file.read()
    original_filename = file.filename or "items.pdf"

    job_id = str(uuid4())
    _set_job(
        job_id,
        created_at=time.time(),
        status="processing",
        message="Старт обработки",
        processed_pages=0,
        total_pages=0,
        filename=(os.path.splitext(original_filename)[0] or "items") + ".xlsx",
    )

    def worker():
        try:
            _set_job(job_id, message="Читаю PDF…")
            parsed_rows, st = parse_items(pdf_bytes)
            output_rows, map_stats = convert_to_output_rows(parsed_rows)

            _set_job(
                job_id,
                processed_pages=int(st.get("processed_pages", 0) or 0),
                total_pages=int(st.get("total_pages", 0) or 0),
                message="Формирую XLSX…",
            )

            if not output_rows:
                _set_job(job_id, status="error", message="Не удалось найти позиции", stats=st)
                return

            st.update(map_stats)
            xlsx_bytes = make_xlsx(output_rows)
            _set_job_result(job_id, xlsx_bytes)
            increment_counter()
            _set_job(job_id, status="done", message="Готово", stats=st)
        except Exception as e:
            _set_job(job_id, status="error", message=f"Ошибка: {e}")

    threading.Thread(target=worker, daemon=True).start()
    return {"job_id": job_id}


@app.get("/job/{job_id}")
def job_status(job_id: str):
    j = _get_job(job_id)
    if not j:
        raise HTTPException(status_code=404, detail="job not found")
    return {
        "status": j.get("status"),
        "message": j.get("message"),
        "processed_pages": int(j.get("processed_pages", 0) or 0),
        "total_pages": int(j.get("total_pages", 0) or 0),
        "stats": j.get("stats"),
        "filename": j.get("filename"),
    }


import urllib.parse


@app.get("/job/{job_id}/download")
def job_download(job_id: str):
    j = _get_job(job_id)
    if not j:
        raise HTTPException(status_code=404, detail="job not found")
    if j.get("status") != "done":
        raise HTTPException(status_code=409, detail="job not done")

    xlsx_bytes = _get_job_result(job_id)
    if not xlsx_bytes:
        raise HTTPException(status_code=404, detail="result not found")

    filename_utf8 = j.get("filename", "items.xlsx")
    filename_ascii = re.sub(r"[^A-Za-z0-9_.-]+", "_", filename_utf8)
    quoted = urllib.parse.quote(filename_utf8)
    headers = {
        "Content-Disposition": (
            f'attachment; filename="{filename_ascii}"; '
            f"filename*=UTF-8''{quoted}"
        )
    }
    return Response(
        content=xlsx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )
